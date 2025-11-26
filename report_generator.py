"""
Модуль для генерации отчетов в Excel
"""
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from datetime import datetime
from typing import List, Dict
import logging

from config import CONNECTION_TYPES

logger = logging.getLogger(__name__)


class ReportGenerator:
    """Класс для генерации Excel-отчетов"""
    
    @staticmethod
    def generate_employee_report(
        employee_name: str,
        connections: List[Dict],
        stats: Dict,
        period_name: str,
        movements: List[Dict] = None
    ) -> str:
        """
        Генерирует Excel-отчет по сотруднику
        
        Args:
            employee_name: ФИО сотрудника
            connections: Список подключений
            stats: Итоговая статистика
            period_name: Название периода
            movements: Список движений материалов и роутеров (опционально)
        
        Returns:
            Путь к созданному файлу
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Отчет"
        
        # Стили
        header_font = Font(name='Arial', size=12, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        title_font = Font(name='Arial', size=14, bold=True)
        title_alignment = Alignment(horizontal='center', vertical='center')
        
        cell_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        number_alignment = Alignment(horizontal='right', vertical='center')
        
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        total_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        total_font = Font(name='Arial', size=11, bold=True)
        
        # Заголовок отчета
        ws.merge_cells('A1:N1')
        ws['A1'] = f"Сводный отчет по монтажнику"
        ws['A1'].font = title_font
        ws['A1'].alignment = title_alignment
        
        # Информация о сотруднике и периоде
        ws.merge_cells('A2:N2')
        ws['A2'] = f"Исполнитель: {employee_name}"
        ws['A2'].font = Font(name='Arial', size=11, bold=True)
        ws['A2'].alignment = cell_alignment
        
        ws.merge_cells('A3:N3')
        ws['A3'] = f"Период: {period_name}"
        ws['A3'].font = Font(name='Arial', size=11)
        ws['A3'].alignment = cell_alignment
        
        ws.merge_cells('A4:N4')
        ws['A4'] = f"Дата формирования: {datetime.now().strftime('%d.%m.%Y %H:%M')}"
        ws['A4'].font = Font(name='Arial', size=10)
        ws['A4'].alignment = cell_alignment
        
        # Заголовки столбцов (строка 6)
        headers = [
            'Столбец',
            'Тип',
            'Исполнители',
            'Адрес подключения',
            'Модель роутера',
            'ВОЛС (всего)',
            'Витая пара (всего)',
            'ВОЛС на исполнителя',
            'Витая пара на исполнителя',
            'SNR бокс',
            'ONU',
            'Медиаконверторы',
            'Связь с подключением',
            'Дата'
        ]
        
        ws.row_dimensions[6].height = 30
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=6, column=col_num)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
        
        # Ширина столбцов
        ws.column_dimensions['A'].width = 12  # Столбец
        ws.column_dimensions['B'].width = 15  # Тип
        ws.column_dimensions['C'].width = 25  # Исполнители
        ws.column_dimensions['D'].width = 30  # Адрес
        ws.column_dimensions['E'].width = 15  # Модель роутера
        ws.column_dimensions['F'].width = 14  # ВОЛС всего
        ws.column_dimensions['G'].width = 14  # Витая пара всего
        ws.column_dimensions['H'].width = 16  # ВОЛС на исполнителя
        ws.column_dimensions['I'].width = 16  # Витая пара на исполнителя
        ws.column_dimensions['J'].width = 18  # SNR бокс
        ws.column_dimensions['K'].width = 18  # ONU
        ws.column_dimensions['L'].width = 22  # МК
        ws.column_dimensions['M'].width = 20  # Связь с подключением
        ws.column_dimensions['N'].width = 18  # Дата
        
        # Данные подключений
        current_row = 7
        for idx, conn in enumerate(connections, 1):
            # Форматируем дату
            try:
                created_at = datetime.fromisoformat(conn['created_at'])
                date_str = created_at.strftime('%d.%m.%Y %H:%M')
            except:
                date_str = conn['created_at']
            
            # Список исполнителей через запятую
            executors = ', '.join(conn['all_employees'])
            
            # Получаем читаемое название типа подключения
            conn_type = conn.get('connection_type', 'mkd')
            type_name = CONNECTION_TYPES.get(conn_type, conn_type)
            
            connection_link = f"Подключение #{conn['id']}" if conn.get('id') else "-"

            row_data = [
                idx,  # Номер по порядку
                type_name,  # Тип подключения
                executors,
                conn['address'],
                conn['router_model'],
                conn.get('total_fiber_meters', conn.get('fiber_meters')),
                conn.get('total_twisted_pair_meters', conn.get('twisted_pair_meters')),
                conn['employee_fiber_meters'],
                conn['employee_twisted_pair_meters'],
                conn.get('snr_spent', conn.get('snr_box_model', '-')),
                conn.get('onu_spent', '-'),
                conn.get('media_spent', '-'),
                connection_link,
                date_str
            ]
            
            for col_num, value in enumerate(row_data, 1):
                cell = ws.cell(row=current_row, column=col_num)
                cell.value = value
                cell.border = border
                
                if col_num in [6, 7, 8, 9]:  # Числовые столбцы
                    cell.alignment = number_alignment
                    cell.number_format = '0.00'
                else:
                    cell.alignment = cell_alignment
            
            current_row += 1
        
        # Итоги
        current_row += 1
        
        # Итого общее
        ws.merge_cells(f'A{current_row}:E{current_row}')
        cell = ws.cell(row=current_row, column=1)
        cell.value = "Итого общее:"
        cell.font = total_font
        cell.fill = total_fill
        cell.alignment = Alignment(horizontal='right', vertical='center')
        cell.border = border
        
        # Итого ВОЛС (всего по подключениям)
        cell = ws.cell(row=current_row, column=6)
        cell.value = stats.get('total_connection_fiber_meters', stats.get('total_fiber_meters', 0))
        cell.font = total_font
        cell.fill = total_fill
        cell.alignment = number_alignment
        cell.number_format = '0.00'
        cell.border = border
        
        # Итого витая пара (всего по подключениям)
        cell = ws.cell(row=current_row, column=7)
        cell.value = stats.get('total_connection_twisted_pair_meters', stats.get('total_twisted_pair_meters', 0))
        cell.font = total_font
        cell.fill = total_fill
        cell.alignment = number_alignment
        cell.number_format = '0.00'
        cell.border = border
        
        # Пустые ячейки для выравнивания
        for col in (8, 9, 10, 11, 12, 13, 14):
            cell = ws.cell(row=current_row, column=col)
            cell.fill = total_fill
            cell.border = border
        
        # Итого для сотрудника (с учетом деления)
        current_row += 1
        ws.merge_cells(f'A{current_row}:E{current_row}')
        cell = ws.cell(row=current_row, column=1)
        cell.value = f"Итого {employee_name}:"
        cell.font = Font(name='Arial', size=12, bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        cell.alignment = Alignment(horizontal='right', vertical='center')
        cell.border = border
        
        # Итого ВОЛС для сотрудника (доля)
        cell = ws.cell(row=current_row, column=8)
        cell.value = stats.get('total_fiber_meters', 0)
        cell.font = Font(name='Arial', size=12, bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        cell.alignment = number_alignment
        cell.number_format = '0.00'
        cell.border = border
        
        # Итого витая пара для сотрудника (доля)
        cell = ws.cell(row=current_row, column=9)
        cell.value = stats.get('total_twisted_pair_meters', 0)
        cell.font = Font(name='Arial', size=12, bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        cell.alignment = number_alignment
        cell.number_format = '0.00'
        cell.border = border
        
        # Пустые ячейки для выравнивания
        for col in (6, 7, 10, 11, 12, 13, 14):
            cell = ws.cell(row=current_row, column=col)
            cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
            cell.border = border
        
        # Создаём второй лист с движениями материалов, если они есть
        if movements and len(movements) > 0:
            ReportGenerator._add_movements_sheet(wb, employee_name, period_name, movements)
        
        # Сохранение файла
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"report_{employee_name.replace(' ', '_')}_{timestamp}.xlsx"
        wb.save(filename)
        
        logger.info(f"Отчет создан: {filename}")
        return filename

    @staticmethod
    def generate_global_report(
        connections: List[Dict],
        stats: Dict,
        period_name: str,
    ) -> str:
        """
        Генерирует Excel-отчет по всем подключениям за период

        Args:
            connections: Список подключений
            stats: Итоговая статистика
            period_name: Название периода

        Returns:
            Путь к созданному файлу
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Отчет"

        header_font = Font(name='Arial', size=12, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        title_font = Font(name='Arial', size=14, bold=True)
        title_alignment = Alignment(horizontal='center', vertical='center')

        cell_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        number_alignment = Alignment(horizontal='right', vertical='center')

        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        total_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        total_font = Font(name='Arial', size=11, bold=True)

        ws.merge_cells('A1:N1')
        ws['A1'] = "Общий сводный отчет по подключениям"
        ws['A1'].font = title_font
        ws['A1'].alignment = title_alignment
        
        ws.merge_cells('A2:N2')
        ws['A2'] = f"Период: {period_name}"
        ws['A2'].font = Font(name='Arial', size=11, bold=True)
        ws['A2'].alignment = cell_alignment
        
        ws.merge_cells('A3:N3')
        ws['A3'] = f"Дата формирования: {datetime.now().strftime('%d.%m.%Y %H:%M')}"
        ws['A3'].font = Font(name='Arial', size=10)
        ws['A3'].alignment = cell_alignment

        headers = [
            'Столбец',
            'Тип',
            'Исполнители',
            'Адрес подключения',
            'Модель роутера',
            'ВОЛС (всего)',
            'Витая пара (всего)',
            'ВОЛС на исполнителя',
            'Витая пара на исполнителя',
            'SNR бокс',
            'ONU',
            'Медиаконверторы',
            'Связь с подключением',
            'Дата'
        ]

        ws.row_dimensions[5].height = 30
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=5, column=col_num)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border

        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 25
        ws.column_dimensions['D'].width = 30
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 14
        ws.column_dimensions['G'].width = 14
        ws.column_dimensions['H'].width = 16
        ws.column_dimensions['I'].width = 16
        ws.column_dimensions['J'].width = 18
        ws.column_dimensions['K'].width = 18
        ws.column_dimensions['L'].width = 22
        ws.column_dimensions['M'].width = 20
        ws.column_dimensions['N'].width = 18

        current_row = 6
        for idx, conn in enumerate(connections, 1):
            try:
                created_at = datetime.fromisoformat(conn['created_at'])
                date_str = created_at.strftime('%d.%m.%Y %H:%M')
            except Exception:
                date_str = conn.get('created_at', '')

            executors = ', '.join(conn.get('all_employees', []))

            conn_type = conn.get('connection_type', 'mkd')
            type_name = CONNECTION_TYPES.get(conn_type, conn_type)

            connection_link = f"Подключение #{conn.get('id')}" if conn.get('id') else "-"

            row_data = [
                idx,
                type_name,
                executors,
                conn.get('address'),
                conn.get('router_model'),
                conn.get('total_fiber_meters', conn.get('fiber_meters', 0)),
                conn.get('total_twisted_pair_meters', conn.get('twisted_pair_meters', 0)),
                conn.get('employee_fiber_meters', 0),
                conn.get('employee_twisted_pair_meters', 0),
                conn.get('snr_spent', conn.get('snr_box_model', '-')),
                conn.get('onu_spent', '-'),
                conn.get('media_spent', '-'),
                connection_link,
                date_str
            ]

            for col_num, value in enumerate(row_data, 1):
                cell = ws.cell(row=current_row, column=col_num)
                cell.value = value
                cell.border = border

                if col_num in [6, 7, 8, 9]:
                    cell.alignment = number_alignment
                    cell.number_format = '0.00'
                else:
                    cell.alignment = cell_alignment

            current_row += 1

        current_row += 1

        ws.merge_cells(f'A{current_row}:E{current_row}')
        cell = ws.cell(row=current_row, column=1)
        cell.value = "Итого:"
        cell.font = total_font
        cell.fill = total_fill
        cell.alignment = Alignment(horizontal='right', vertical='center')
        cell.border = border

        cell = ws.cell(row=current_row, column=6)
        cell.value = stats.get('total_connection_fiber_meters', stats.get('total_fiber_meters', 0))
        cell.font = total_font
        cell.fill = total_fill
        cell.alignment = number_alignment
        cell.number_format = '0.00'
        cell.border = border
    
        cell = ws.cell(row=current_row, column=7)
        cell.value = stats.get('total_connection_twisted_pair_meters', stats.get('total_twisted_pair_meters', 0))
        cell.font = total_font
        cell.fill = total_fill
        cell.alignment = number_alignment
        cell.number_format = '0.00'
        cell.border = border

        cell = ws.cell(row=current_row, column=8)
        cell.value = stats.get('total_fiber_meters', 0)
        cell.font = total_font
        cell.fill = total_fill
        cell.alignment = number_alignment
        cell.number_format = '0.00'
        cell.border = border

        cell = ws.cell(row=current_row, column=9)
        cell.value = stats.get('total_twisted_pair_meters', 0)
        cell.font = total_font
        cell.fill = total_fill
        cell.alignment = number_alignment
        cell.number_format = '0.00'
        cell.border = border

        file_name = f"global_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        wb.save(file_name)
        logger.info("Общий отчет создан: %s", file_name)
        return file_name
    
    @staticmethod
    def _add_movements_sheet(wb: Workbook, employee_name: str, period_name: str, movements: List[Dict]):
        """
        Добавляет лист с движениями материалов и роутеров
        
        Args:
            wb: Workbook объект
            employee_name: ФИО сотрудника
            period_name: Название периода
            movements: Список движений
        """
        # Создаём новый лист
        ws = wb.create_sheet(title="Движение материалов")
        
        # Стили
        header_font = Font(name='Arial', size=12, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        title_font = Font(name='Arial', size=14, bold=True)
        title_alignment = Alignment(horizontal='center', vertical='center')
        
        cell_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        number_alignment = Alignment(horizontal='right', vertical='center')
        
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Зелёный для добавления
        add_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        # Красный для списания
        deduct_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
        
        # Заголовок
        ws.merge_cells('A1:G1')
        ws['A1'] = f"Движение материалов и роутеров"
        ws['A1'].font = title_font
        ws['A1'].alignment = title_alignment
        
        # Информация
        ws.merge_cells('A2:G2')
        ws['A2'] = f"Исполнитель: {employee_name}"
        ws['A2'].font = Font(name='Arial', size=11, bold=True)
        ws['A2'].alignment = cell_alignment
        
        ws.merge_cells('A3:G3')
        ws['A3'] = f"Период: {period_name}"
        ws['A3'].font = Font(name='Arial', size=11)
        ws['A3'].alignment = cell_alignment
        
        # Заголовки столбцов
        headers = [
            'Дата',
            'Операция',
            'Тип',
            'Название',
            'Количество',
            'Остаток',
            'Связь с подключением'
        ]
        
        ws.row_dimensions[5].height = 30
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=5, column=col_num)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
        
        # Ширина столбцов
        ws.column_dimensions['A'].width = 18  # Дата
        ws.column_dimensions['B'].width = 12  # Операция
        ws.column_dimensions['C'].width = 15  # Тип
        ws.column_dimensions['D'].width = 20  # Название
        ws.column_dimensions['E'].width = 12  # Количество
        ws.column_dimensions['F'].width = 12  # Остаток
        ws.column_dimensions['G'].width = 20  # Связь
        
        # Данные движений
        current_row = 6
        for mov in movements:
            # Форматируем дату
            try:
                created_at = datetime.fromisoformat(mov['created_at'])
                date_str = created_at.strftime('%d.%m.%Y %H:%M')
            except:
                date_str = mov['created_at']
            
            # Операция
            operation = "Добавление" if mov['operation_type'] == 'add' else "Списание"
            
            # Тип
            type_map = {
                'fiber': 'ВОЛС',
                'twisted_pair': 'Витая пара',
                'router': 'Роутер',
                'snr_box': 'SNR бокс',
                'onu': 'ONU',
                'media_converter': 'Медиаконвертор'
            }
            item_type = type_map.get(mov['item_type'], mov['item_type'])
            
            # Количество
            if mov['item_type'] in ('router', 'snr_box', 'onu', 'media_converter'):
                quantity_str = f"{int(mov['quantity'])} шт."
                balance_str = f"{int(mov['balance_after'])} шт."
            else:
                quantity_str = f"{mov['quantity']} м"
                balance_str = f"{mov['balance_after']} м"
            
            # Связь с подключением
            conn_link = f"Подключение #{mov['connection_id']}" if mov['connection_id'] else "-"
            
            row_data = [
                date_str,
                operation,
                item_type,
                mov['item_name'],
                quantity_str,
                balance_str,
                conn_link
            ]
            
            # Определяем цвет фона
            row_fill = add_fill if mov['operation_type'] == 'add' else deduct_fill
            
            for col_num, value in enumerate(row_data, 1):
                cell = ws.cell(row=current_row, column=col_num)
                cell.value = value
                cell.border = border
                cell.fill = row_fill
                
                if col_num in [5, 6]:  # Количество и остаток
                    cell.alignment = number_alignment
                else:
                    cell.alignment = cell_alignment
            
            current_row += 1
        
        logger.info(f"Добавлен лист 'Движение материалов' с {len(movements)} записями")
